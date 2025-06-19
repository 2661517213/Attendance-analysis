import psycopg2
import pandas as pd
import re
from datetime import datetime
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from holidays import HOLIDAYS, MONTH, get_working_days
import os
from config import DB_CONFIG
import sys

# 强制刷新输出缓冲区
def flush_print(*args, **kwargs):
    """带缓冲刷新的print函数"""
    print(*args, **kwargs)
    sys.stdout.flush()

# 创建输出目录
OUTPUT_DIR = "output"
os.makedirs(OUTPUT_DIR, exist_ok=True)

# 生成带时间戳的输出文件名
timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
output_file = os.path.join(OUTPUT_DIR, f"考勤明细及统计_{timestamp}.xlsx")

def get_db_connection():
    """创建数据库连接"""
    return psycopg2.connect(**DB_CONFIG)

def count_attendance_status(row):
    """统计考勤状态，仅对请假排除休息日"""
    counts = {
        "正常次数": 0,
        "迟到次数": 0,
        "早退次数": 0,
        "缺卡次数": 0,
        "旷工次数": 0,
        "出差次数": 0,
        "请假次数": 0,
        "钉钉加班时长(h)": 0.0,
        "飞书加班时长(h)": 0.0,
        "总加班时长(h)": 0.0  # 新增合并加班时长
    }
    
    for day in range(1, 32):
        day_col = f"第{day}天"
        if day_col not in row:
            continue
            
        status = str(row[day_col])
        if not status or status == 'nan':
            continue
        
        # 获取日期字符串
        day_str = f"{day:02d}"
        
        # 统计所有状态（除请假外）
        if "正常" in status:
            counts["正常次数"] += 1
        if "迟到" in status:
            counts["迟到次数"] += 1
        if "早退" in status:
            counts["早退次数"] += 1
        if "缺卡" in status:
            counts["缺卡次数"] += 1
        if "旷工" in status:
            counts["旷工次数"] += 1
        if "出差" in status:
            counts["出差次数"] += 1
            
        # 请假只在非休息日统计
        if "请假" in status and day_str not in HOLIDAYS:
            counts["请假次数"] += 1
            
        overtime_matches = re.findall(r'(钉钉加班|飞书加班)\((\d+\.?\d*)h\)', status)
        for source, hours in overtime_matches:
            try:
                hours_float = float(hours)
                if source == "钉钉加班":
                    counts["钉钉加班时长(h)"] += hours_float
                    counts["总加班时长(h)"] += hours_float  # 累加到总时长
                elif source == "飞书加班":
                    counts["飞书加班时长(h)"] += hours_float
                    counts["总加班时长(h)"] += hours_float  # 累加到总时长
            except ValueError:
                flush_print(f"警告: 无法转换加班时长 '{hours}' 为数字")
                continue
    return pd.Series(counts)

def format_attendance_status(status, day):
    """格式化考勤状态，为休息日和其他状态添加标记"""
    if not status or str(status) == 'nan':
        return ''
        
    status = str(status)
    formatted = []
    
    # 检查是否为休息日
    day_str = f"{day:02d}"
    is_holiday = day_str in HOLIDAYS
    
    # 如果是休息日，添加休息日标记
    if is_holiday:
        formatted.append("🏠")
    
    # 添加其他状态的图标
    if "正常" in status:
        formatted.append("✅")
    if "迟到" in status:
        formatted.append("⏰")
    if "早退" in status:
        formatted.append("⚡")
    if "缺卡" in status:
        formatted.append("❌")
    if "旷工" in status:
        formatted.append("⛔")
    if "出差" in status:
        formatted.append("🚗")
    if "请假" in status:
        formatted.append("📝")
    
    # 组织显示格式
    if formatted:
        if is_holiday:
            return f"{''.join(formatted)} 休息日\n{status}"
        else:
            return f"{''.join(formatted)} {status}"
    return status

def analyze_attendance():
    """分析考勤数据并添加统计结果"""
    conn = get_db_connection()
    try:
        query = """
        SELECT * FROM attendance_result
        """
        df = pd.read_sql_query(query, conn)
        
        # 计算统计结果并添加到原DataFrame
        statistics = df.apply(count_attendance_status, axis=1)
        df = pd.concat([df, statistics], axis=1)
        
        # 在导出到Excel之前，格式化考勤状态
        for day in range(1, 32):
            day_col = f"第{day}天"
            if day_col in df.columns:
                # 传入当前日期以判断是否为休息日
                df[day_col] = df.apply(lambda row: format_attendance_status(row[day_col], day), axis=1)
        
        # 添加应出勤天数列
        working_days = get_working_days()
        df['应出勤天数'] = int(working_days)
        
        # 确保缺卡次数为整数
        df['缺卡次数'] = df['缺卡次数'].astype(int)
        
        # 计算实际出勤天数（应出勤天数减去缺卡和请假次数）
        df['实际出勤天数'] = df['应出勤天数'] - df['缺卡次数']
        
        # 修改统计列定义，添加实际出勤天数列
        stat_columns = {
            38: ('应出勤天数', 12),
            39: ('实际出勤天数', 12),
            40: ('正常次数', 12),
            41: ('迟到次数', 12),
            42: ('早退次数', 12),
            43: ('缺卡次数', 12),
            44: ('旷工次数', 12),
            45: ('出差次数', 12),
            46: ('请假次数', 12),
            47: ('钉钉加班时长(h)', 15),
            48: ('飞书加班时长(h)', 15),
            49: ('总加班时长(h)', 15)  
        }
        
        # 导出到Excel时设置更好的格式
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='考勤明细')
            
            # 获取工作簿和工作表
            workbook = writer.book
            worksheet = writer.sheets['考勤明细']
            
            # 导入所需的样式
            from openpyxl.utils import get_column_letter
            from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
            
            # 设置表头样式
            header_font = Font(name='微软雅黑', bold=True, size=11, color='000000')
            header_fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            # 应用表头样式
            for cell in worksheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.border = border
                cell.alignment = header_alignment
            
            # 冻结首行
            worksheet.freeze_panes = 'B2'
        
            
            # 设置列宽
            base_columns = {
                'A': ('姓名', 15),
                'B': ('考勤组', 20),
                'C': ('部门', 25),
                'D': ('工号', 12),
                'E': ('职位', 15),
                'F': ('UserId', 20)
            }
            
            # 设置基础列的宽度
            for col, (_, width) in base_columns.items():
                worksheet.column_dimensions[col].width = width
                
            # 设置每日考勤列宽度（第7列到第37列）
            for i in range(7, 38):
                col_letter = get_column_letter(i)
                worksheet.column_dimensions[col_letter].width = 40  # 增加宽度
                
            # 应用统计列设置
            for col_num, (col_name, width) in stat_columns.items():
                col_letter = get_column_letter(col_num)
                worksheet.column_dimensions[col_letter].width = width
            
            # 设置数据区域的对齐方式
            data_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            for row in worksheet.iter_rows(min_row=2):  # 从第二行开始
                for cell in row:
                    cell.alignment = data_alignment
                    cell.border = border
        
        flush_print(f"✅ 考勤明细及统计已导出到: {output_file}")
        
        # 打印总体统计信息
        flush_print("\n📊 考勤统计概览:")
        flush_print(f"总人数: {len(df)}")
        flush_print(f"应出勤天数: {working_days}")
        flush_print(f"平均实际出勤天数: {df['实际出勤天数'].mean():.1f}")
        flush_print("\n各类情况汇总:")
        
        # 打印统计信息时添加实际出勤天数
        stat_columns = [
            "正常次数", 
            "迟到次数", 
            "早退次数", 
            "缺卡次数", 
            "旷工次数", 
            "出差次数", 
            "请假次数"  # 移除了请假天数
        ]
        
        # 打印统计信息
        for column in stat_columns:
            try:
                total = df[column].sum()
                flush_print(f"{column}: {total:.1f}")
            except KeyError as e:
                flush_print(f"警告: 未找到列 '{column}'")
    
        return df
        
    except Exception as e:
        flush_print(f"❌ 处理出错: {e}")
        return None
    finally:
        conn.close()

def main():
    # 分析并导出考勤数据
    summary_df = analyze_attendance()
    
    if summary_df is not None:
        # 打印总体统计信息
        flush_print("\n✅ 统计完成")

if __name__ == "__main__":
    main()